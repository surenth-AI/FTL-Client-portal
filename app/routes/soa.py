from flask import Blueprint, render_template, request, flash, redirect, url_for, session, send_file
from flask_login import login_required, current_user
from functools import wraps
import pandas as pd
import os
import json
import uuid
import io
from werkzeug.utils import secure_filename
from app.models.models import db, VendorInvoice, Job
from openpyxl.styles import Font, PatternFill, Alignment

soa_bp = Blueprint('soa', __name__, url_prefix='/soa')



@soa_bp.route('/', methods=['GET'])
@login_required
def index():
    return render_template('soa/index.html')

@soa_bp.route('/upload', methods=['POST'])
@login_required
def upload():
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('soa.index'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('soa.index'))

    if file and (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        try:
            # Save the file temporarily
            filename = secure_filename(file.filename)
            upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            filepath = os.path.join(upload_dir, filename)
            file.save(filepath)

            # Process the SOA
            engine = 'xlrd' if file.filename.endswith('.xls') else 'openpyxl'
            df = pd.read_excel(filepath, engine=engine)
            
            # Recreate DataFrame using row 9 as header
            headers = df.iloc[9].fillna('').astype(str).tolist()
            df = df.iloc[10:].reset_index(drop=True)
            df.columns = headers

            mismatches = []
            matched = []
            orphan_jobs = []
            
            # Map system invoices (only for current user)
            system_invoices = VendorInvoice.query.filter_by(user_id=current_user.id).all()
            system_invoice_map = {inv.invoice_number: inv for inv in system_invoices if inv.invoice_number}
            
            # And Jobs
            system_jobs = Job.query.filter_by(user_id=current_user.id).all()
            system_job_map = {job.job_number: job for job in system_jobs}

            soa_job_numbers = set()

            for _, row in df.iterrows():
                job_no = str(row.get('Job No.', '')).strip()
                if not job_no or job_no == 'nan':
                    continue
                    
                soa_job_numbers.add(job_no)
                
                # Fetch debit/credit
                try:
                    debit = float(row.get('Debit', 0) or 0)
                except: debit = 0.0
                
                try:
                    credit = float(row.get('Credit', 0) or 0)
                except: credit = 0.0
                
                soa_amount = abs(debit + credit) # Taking the net amount
                
                # Compare against system invoice / job
                # Priority: Check if invoice exists for this Job No.
                inv = system_invoice_map.get(job_no)
                
                if inv:
                    # Validate amounts
                    if abs(inv.amount - soa_amount) > 0.01:
                        mismatches.append({
                            'job_no': job_no,
                            'system_amount': inv.amount,
                            'soa_amount': soa_amount,
                            'diff': abs(inv.amount - soa_amount)
                        })
                    else:
                        matched.append({
                            'job_no': job_no,
                            'amount': soa_amount,
                            'status': 'Perfect Match'
                        })
                else:
                    # Invoice doesn't exist, check if job exists
                    if job_no in system_job_map:
                        # Job exists, but no invoice generated yet!
                        mismatches.append({
                            'job_no': job_no,
                            'system_amount': 0.0,
                            'soa_amount': soa_amount,
                            'diff': soa_amount,
                            'note': 'No Invoice found for this Job'
                        })
                    else:
                        orphan_jobs.append({
                            'job_no': job_no,
                            'soa_amount': soa_amount
                        })

            # Identify Missing Invoices (System invoices that don't appear in the SOA)
            missing_invoices = []
            for inv in system_invoices:
                if inv.invoice_number and inv.invoice_number not in soa_job_numbers:
                    missing_invoices.append({
                        'job_no': inv.invoice_number,
                        'amount': inv.amount,
                        'supplier': inv.supplier
                    })

            # Cleanup file
            if os.path.exists(filepath):
                os.remove(filepath)
                
            # Persist results for export
            result_id = str(uuid.uuid4())
            results_data = {
                'matched': matched,
                'mismatches': mismatches,
                'missing_invoices': missing_invoices,
                'orphan_jobs': orphan_jobs,
                'summary': {
                    'matched_count': len(matched),
                    'mismatch_count': len(mismatches),
                    'missing_count': len(missing_invoices),
                    'orphan_count': len(orphan_jobs)
                }
            }
            results_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads', 'soa_results')
            os.makedirs(results_dir, exist_ok=True)
            results_path = os.path.join(results_dir, f"{result_id}.json")
            with open(results_path, 'w') as f:
                json.dump(results_data, f)

            flash(f"Processed SOA successfully! {len(matched)} exact matches found.", "success")
            return render_template('soa/results.html', 
                                   mismatches=mismatches, 
                                   missing_invoices=missing_invoices,
                                   orphan_jobs=orphan_jobs,
                                   result_id=result_id)

        except Exception as e:
            flash(f"Error parsing Statement of Accounts: {str(e)}", "error")
            return redirect(url_for('soa.index'))
    else:
        flash("Invalid file format. Please upload .xls or .xlsx files.", "error")
        return redirect(url_for('soa.index'))

@soa_bp.route('/export/<result_id>')
@login_required
def export(result_id):
    results_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads', 'soa_results', f"{result_id}.json")
    if not os.path.exists(results_path):
        flash("Export data not found or expired.", "error")
        return redirect(url_for('soa.index'))
    
    with open(results_path, 'r') as f:
        data = json.load(f)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Exact Matches
        df_exact = pd.DataFrame(data.get('matched', []))
        if not df_exact.empty:
            df_exact = df_exact.rename(columns={'job_no': 'Ref No / Job ID', 'amount': 'Matched Amount', 'status': 'Status'})
        else:
            df_exact = pd.DataFrame(columns=['Ref No / Job ID', 'Matched Amount', 'Status'])
        df_exact.to_excel(writer, sheet_name='Exact Matches', index=False)
        
        # Sheet 2: Price Mismatches
        df_m = pd.DataFrame(data.get('mismatches', []))
        if not df_m.empty:
            column_mapping = {
                'job_no': 'Ref No / Job ID',
                'system_amount': 'System Amount',
                'soa_amount': 'SOA Amount',
                'diff': 'Variance',
                'note': 'Audit Note'
            }
            df_m = df_m.rename(columns=column_mapping)
        else:
            df_m = pd.DataFrame(columns=['Ref No / Job ID', 'System Amount', 'SOA Amount', 'Variance', 'Audit Note'])
        df_m.to_excel(writer, sheet_name='Price Mismatches', index=False)
        
        # Sheet 3: Missing from SOA
        df_mi = pd.DataFrame(data.get('missing_invoices', []))
        if not df_mi.empty:
            df_mi = df_mi.rename(columns={'job_no': 'Ref No', 'amount': 'System Amount', 'supplier': 'Supplier'})
        else:
            df_mi = pd.DataFrame(columns=['Ref No', 'System Amount', 'Supplier'])
        df_mi.to_excel(writer, sheet_name='Missing from SOA', index=False)
            
        # Sheet 4: Orphan Records
        df_oj = pd.DataFrame(data.get('orphan_jobs', []))
        if not df_oj.empty:
            df_oj = df_oj.rename(columns={'job_no': 'Ref No', 'soa_amount': 'SOA Amount'})
        else:
            df_oj = pd.DataFrame(columns=['Ref No', 'SOA Amount'])
        df_oj.to_excel(writer, sheet_name='Orphan Records', index=False)

        # Premium Styling
        for sheetname in writer.sheets:
            sheet = writer.sheets[sheetname]
            
            # Header Styling: Dark Blue / Slate with white text
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Alternate row coloring for clarity
            row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = row_fill
            
            # Borders and Padding (simulated width)
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = min(max_length + 6, 60)

    output.seek(0)
    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, 
                     download_name=f"SOA_Reconciliation_Detailed_{result_id[:8]}.xlsx")


